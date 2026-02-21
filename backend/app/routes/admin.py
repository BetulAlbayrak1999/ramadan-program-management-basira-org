import io
from datetime import date, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, Response
from sqlalchemy import or_
from sqlalchemy.orm import Session
from app.database import get_db
from app.config import settings as app_settings
from app.models.user import User
from app.models.daily_card import DailyCard
from app.models.halqa import Halqa
from app.dependencies import RoleChecker
from typing import List
from pydantic import BaseModel
from app.schemas.user import (
    AdminUserUpdate, AdminResetPassword, SetRole,
    AssignHalqa, RejectRegistration, user_to_response,
)
from sqlalchemy import func
from app.schemas.halqa import HalqaCreate, HalqaUpdate, AssignMembers, halqa_to_response
from app.schemas.daily_card import card_to_response

router = APIRouter(prefix="/api/admin", tags=["admin"])

require_admin = RoleChecker("super_admin")


# ─── User Management ──────────────────────────────────────────────────────────


@router.get("/registrations")
def get_registrations(
    status: str = Query("pending"),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get all pending registrations."""
    if status == "all":
        users = db.query(User).order_by(User.created_at.desc()).all()
    else:
        users = db.query(User).filter_by(status=status).order_by(User.created_at.desc()).all()
    return {"users": [user_to_response(u) for u in users]}


@router.post("/registration/{user_id}/approve")
def approve_registration(
    user_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Approve a registration request."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")

    user.status = "active"
    user.rejection_note = None
    db.commit()
    db.refresh(user)
    return {"message": "تم قبول الطلب", "user": user_to_response(user)}


@router.post("/registration/{user_id}/reject")
def reject_registration(
    user_id: int,
    data: RejectRegistration = None,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Reject a registration request."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")

    user.status = "rejected"
    user.rejection_note = data.note if data else ""
    db.commit()
    db.refresh(user)
    return {"message": "تم رفض الطلب", "user": user_to_response(user)}


@router.get("/users")
def get_all_users(
    status: str = Query(None),
    gender: str = Query(None),
    halqa_id: int = Query(None),
    search: str = Query(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get all users with optional filters."""
    query = db.query(User)

    if status:
        query = query.filter_by(status=status)
    if gender:
        query = query.filter_by(gender=gender)
    if halqa_id:
        query = query.filter_by(halqa_id=halqa_id)
    if search:
        query = query.filter(
            or_(
                User.full_name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
            )
        )

    users = query.order_by(User.created_at.desc()).all()
    return {"users": [user_to_response(u) for u in users]}


@router.get("/user/{user_id}")
def get_user(
    user_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get user details."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")
    return {"user": user_to_response(user)}


@router.put("/user/{user_id}")
def update_user(
    user_id: int,
    data: AdminUserUpdate,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Update user details."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")

    allowed = ["full_name", "gender", "age", "phone", "country", "referral_source", "status", "halqa_id"]
    for field in allowed:
        value = getattr(data, field, None)
        if value is not None:
            setattr(user, field, value)

    db.commit()
    db.refresh(user)
    return {"message": "تم تحديث البيانات", "user": user_to_response(user)}


@router.post("/user/{user_id}/reset-password")
def admin_reset_password(
    user_id: int,
    data: AdminResetPassword,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Reset user password by admin."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")

    user.set_password(data.new_password)
    db.commit()
    return {"message": "تم إعادة تعيين كلمة المرور"}


@router.post("/user/{user_id}/withdraw")
def withdraw_user(
    user_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Mark user as withdrawn."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")

    user.status = "withdrawn"
    db.commit()
    db.refresh(user)
    return {"message": "تم سحب المشارك", "user": user_to_response(user)}


@router.post("/user/{user_id}/activate")
def activate_user(
    user_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Re-activate user."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")

    user.status = "active"
    db.commit()
    db.refresh(user)
    return {"message": "تم تفعيل المشارك", "user": user_to_response(user)}


# ─── Role Management ──────────────────────────────────────────────────────────


@router.post("/user/{user_id}/set-role")
def set_user_role(
    user_id: int,
    data: SetRole,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Set user role (supervisor, super_admin, participant)."""
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(404, detail="المستخدم غير موجود")

    if data.role not in ["participant", "supervisor", "super_admin"]:
        raise HTTPException(400, detail="الصلاحية غير صالحة")

    # Only primary admin can manage super_admin roles
    primary_email = app_settings.SUPER_ADMIN_EMAIL.lower()
    if data.role == "super_admin" or target.role == "super_admin":
        if admin.email != primary_email:
            raise HTTPException(403, detail="فقط المشرف الرئيسي يمكنه إدارة صلاحيات السوبر آدمن")

    target.role = data.role
    db.commit()
    db.refresh(target)
    return {"message": "تم تحديث الصلاحية", "user": user_to_response(target)}


# ─── Halqa Management ─────────────────────────────────────────────────────────


@router.get("/halqas")
def get_halqas(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get all halqas."""
    halqas = db.query(Halqa).all()
    return {"halqas": [halqa_to_response(h) for h in halqas]}


@router.post("/halqa")
def create_halqa(
    data: HalqaCreate,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Create a new halqa."""
    name = data.name.strip()
    if not name:
        raise HTTPException(400, detail="اسم الحلقة مطلوب")

    if db.query(Halqa).filter_by(name=name).first():
        raise HTTPException(400, detail="اسم الحلقة موجود مسبقاً")

    halqa = Halqa(name=name, supervisor_id=data.supervisor_id)
    db.add(halqa)
    db.commit()
    db.refresh(halqa)
    return {"message": "تم إنشاء الحلقة", "halqa": halqa_to_response(halqa)}


@router.put("/halqa/{halqa_id}")
def update_halqa(
    halqa_id: int,
    data: HalqaUpdate,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Update halqa details."""
    halqa = db.get(Halqa, halqa_id)
    if not halqa:
        raise HTTPException(404, detail="الحلقة غير موجودة")

    if data.name is not None:
        halqa.name = data.name
    if data.supervisor_id is not None:
        halqa.supervisor_id = data.supervisor_id

    db.commit()
    db.refresh(halqa)
    return {"message": "تم تحديث الحلقة", "halqa": halqa_to_response(halqa)}


@router.post("/halqa/{halqa_id}/assign-members")
def assign_members_to_halqa(
    halqa_id: int,
    data: AssignMembers,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Assign members to a halqa."""
    halqa = db.get(Halqa, halqa_id)
    if not halqa:
        raise HTTPException(404, detail="الحلقة غير موجودة")

    for uid in data.user_ids:
        user = db.get(User, uid)
        if user:
            user.halqa_id = halqa_id

    db.commit()
    return {"message": "تم تعيين المشاركين"}


@router.post("/user/{user_id}/assign-halqa")
def assign_user_halqa(
    user_id: int,
    data: AssignHalqa,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Assign a single user to a halqa."""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, detail="المستخدم غير موجود")

    user.halqa_id = data.halqa_id
    db.commit()
    db.refresh(user)
    return {"message": "تم تعيين الحلقة", "user": user_to_response(user)}


# ─── Analytics Dashboard ──────────────────────────────────────────────────────


def _build_analytics_results(
    db: Session,
    gender: str = None,
    halqa_id: int = None,
    supervisor: str = None,
    member: str = None,
    min_pct: float = None,
    max_pct: float = None,
    period: str = "all",
    date_from: str = None,
    date_to: str = None,
    sort_by: str = "score",
    sort_order: str = "desc",
):
    """Shared helper for analytics and export."""
    query = db.query(User).filter_by(status="active")

    if gender:
        query = query.filter_by(gender=gender)
    if halqa_id:
        query = query.filter_by(halqa_id=halqa_id)
    if member:
        query = query.filter(User.full_name.ilike(f"%{member}%"))
    if supervisor:
        halqas = db.query(Halqa).join(User, Halqa.supervisor_id == User.id).filter(
            User.full_name.ilike(f"%{supervisor}%")
        ).all()
        halqa_ids = [h.id for h in halqas]
        if halqa_ids:
            query = query.filter(User.halqa_id.in_(halqa_ids))

    users = query.all()

    # Date range
    today = date.today()
    start_date = None
    end_date = None

    if date_from:
        start_date = date.fromisoformat(date_from)
    elif period == "weekly":
        start_date = today - timedelta(days=today.weekday())
    elif period == "monthly":
        start_date = today.replace(day=1)

    if date_to:
        end_date = date.fromisoformat(date_to)

    # Calculate max based on total days in range (not just submitted cards)
    max_per_day = len(DailyCard.SCORE_FIELDS) * 10  # 110
    total_days = None
    if start_date:
        range_end = end_date or today
        total_days = (range_end - start_date).days + 1

    results = []
    for u in users:
        card_query = db.query(DailyCard).filter_by(user_id=u.id)
        if start_date:
            card_query = card_query.filter(DailyCard.date >= start_date)
        if end_date:
            card_query = card_query.filter(DailyCard.date <= end_date)

        cards = card_query.all()
        total = sum(c.total_score for c in cards)
        if total_days:
            max_total = total_days * max_per_day
        else:
            max_total = sum(c.max_score for c in cards) if cards else 0
        pct = round((total / max_total) * 100, 1) if max_total > 0 else 0

        if min_pct is not None and pct < min_pct:
            continue
        if max_pct is not None and pct > max_pct:
            continue

        results.append({
            "user_id": u.id,
            "member_id": u.member_id,
            "full_name": u.full_name,
            "gender": u.gender,
            "halqa_name": u.halqa.name if u.halqa else "بدون حلقة",
            "supervisor_name": u.halqa.supervisor.full_name if u.halqa and u.halqa.supervisor else "-",
            "total_score": total,
            "max_score": max_total,
            "percentage": pct,
            "cards_count": len(cards),
        })

    if sort_by == "name":
        results.sort(key=lambda x: x["full_name"], reverse=(sort_order == "desc"))
    else:
        results.sort(key=lambda x: x["total_score"], reverse=(sort_order == "desc"))

    for i, r in enumerate(results):
        r["rank"] = i + 1

    return results


@router.get("/analytics")
def get_analytics(
    gender: str = Query(None),
    halqa_id: int = Query(None),
    supervisor: str = Query(None),
    member: str = Query(None),
    min_pct: float = Query(None),
    max_pct: float = Query(None),
    period: str = Query("all"),
    date_from: str = Query(None),
    date_to: str = Query(None),
    sort_by: str = Query("score"),
    sort_order: str = Query("desc"),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get comprehensive analytics."""
    results = _build_analytics_results(
        db, gender=gender, halqa_id=halqa_id, supervisor=supervisor,
        member=member, min_pct=min_pct, max_pct=max_pct, period=period,
        date_from=date_from, date_to=date_to, sort_by=sort_by, sort_order=sort_order,
    )

    total_active = db.query(User).filter_by(status="active").count()
    total_pending = db.query(User).filter_by(status="pending").count()
    total_halqas = db.query(Halqa).count()

    return {
        "results": results,
        "summary": {
            "total_active": total_active,
            "total_pending": total_pending,
            "total_halqas": total_halqas,
            "filtered_count": len(results),
        },
    }


@router.get("/user/{user_id}/cards")
def get_user_cards(
    user_id: int,
    date_from: str = Query(None),
    date_to: str = Query(None),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get all daily cards for a specific user, with optional date filtering."""
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(404, detail="المستخدم غير موجود")

    card_query = db.query(DailyCard).filter_by(user_id=user_id)
    if date_from:
        card_query = card_query.filter(DailyCard.date >= date.fromisoformat(date_from))
    if date_to:
        card_query = card_query.filter(DailyCard.date <= date.fromisoformat(date_to))

    cards = card_query.order_by(DailyCard.date.desc()).all()
    return {
        "member": user_to_response(target),
        "cards": [card_to_response(c) for c in cards],
    }


# ─── Import / Export ──────────────────────────────────────────────────────────


@router.get("/export")
def export_data(
    format: str = Query("csv"),
    gender: str = Query(None),
    halqa_id: int = Query(None),
    supervisor: str = Query(None),
    member: str = Query(None),
    min_pct: float = Query(None),
    max_pct: float = Query(None),
    period: str = Query("all"),
    date_from: str = Query(None),
    date_to: str = Query(None),
    sort_by: str = Query("score"),
    sort_order: str = Query("desc"),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Export analytics data as CSV or XLSX with all applied filters."""
    import csv
    from openpyxl import Workbook

    results = _build_analytics_results(
        db, gender=gender, halqa_id=halqa_id, supervisor=supervisor,
        member=member, min_pct=min_pct, max_pct=max_pct, period=period,
        date_from=date_from, date_to=date_to, sort_by=sort_by, sort_order=sort_order,
    )

    gender_map = {"male": "ذكر", "female": "أنثى"}
    rows = []
    for r in results:
        rows.append({
            "الترتيب": r["rank"],
            "رقم العضوية": r["member_id"],
            "الاسم": r["full_name"],
            "الجنس": gender_map.get(r["gender"], r["gender"]),
            "الحلقة": r["halqa_name"],
            "المشرف": r["supervisor_name"],
            "مجموع النقاط": r["total_score"],
            "الحد الأعلى": r["max_score"],
            "عدد البطاقات": r["cards_count"],
            "النسبة %": r["percentage"],
        })

    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "النتائج"
        ws.sheet_view.rightToLeft = True

        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row[h] for h in headers])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ramadan_results.xlsx"},
        )
    else:
        # UTF-8 BOM so Excel opens Arabic correctly
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        csv_bytes = "\ufeff" + output.getvalue()
        return Response(
            content=csv_bytes.encode("utf-8"),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": "attachment; filename=ramadan_results.csv"},
        )


@router.post("/import")
def import_users(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Import users from Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(file.file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required_headers = ["الاسم", "الجنس", "العمر", "الهاتف", "البريد", "الدولة"]

    for h in required_headers:
        if h not in headers:
            raise HTTPException(400, detail=f"العمود {h} مفقود من الملف")

    max_mid = db.query(func.max(User.member_id)).scalar()
    next_member_id = (max_mid + 1) if max_mid else 1000

    imported = 0
    errors = []
    seen_emails = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_data = dict(zip(headers, row))
        try:
            raw_email = row_data.get("البريد")
            email = str(raw_email).lower().strip() if raw_email is not None else ""
            if not email or email == "none":
                errors.append(f"صف {row_idx}: البريد فارغ")
                continue
            if email in seen_emails:
                errors.append(f"صف {row_idx}: بريد مكرر في الملف")
                continue
            if db.query(User).filter_by(email=email).first():
                errors.append(f"صف {row_idx}: البريد مسجل مسبقاً ({email})")
                continue

            seen_emails.add(email)
            raw_gender = str(row_data.get("الجنس", "")).strip()
            gender_map = {"ذكر": "male", "أنثى": "female", "male": "male", "female": "female"}
            gender = gender_map.get(raw_gender, raw_gender)

            raw_age = row_data.get("العمر", 0)
            age = int(raw_age) if raw_age is not None and str(raw_age).strip() else 0

            user = User(
                member_id=next_member_id,
                full_name=str(row_data.get("الاسم") or "").strip(),
                gender=gender,
                age=age,
                phone=str(row_data.get("الهاتف") or "").strip(),
                email=email,
                country=str(row_data.get("الدولة") or "").strip(),
                referral_source=str(row_data.get("المصدر") or "").strip(),
                status="pending",
                role="participant",
            )
            user.set_password("123456")  # Default password
            db.add(user)
            db.flush()
            next_member_id += 1
            imported += 1
        except Exception as e:
            db.rollback()
            errors.append(f"صف {row_idx}: {str(e)}")

    db.commit()
    return {
        "message": f"تم استيراد {imported} مشارك في قائمة الانتظار",
        "errors": errors,
    }


@router.get("/import-template")
def get_import_template(
    admin: User = Depends(require_admin),
):
    """Download import template."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "قالب الاستيراد"
    ws.sheet_view.rightToLeft = True

    headers = ["الاسم", "الجنس", "العمر", "الهاتف", "البريد", "الدولة", "المصدر"]
    ws.append(headers)

    # Example row
    ws.append(["أحمد محمد علي", "ذكر", 25, "+966500000000", "ahmed@example.com", "السعودية", "صديق"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )


# ─── Bulk Actions ────────────────────────────────────────────────────────────


class BulkUserIds(BaseModel):
    user_ids: List[int]


class BulkAssignHalqa(BaseModel):
    user_ids: List[int]
    halqa_id: int | None = None


@router.post("/bulk/approve")
def bulk_approve(
    data: BulkUserIds,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    count = 0
    for uid in data.user_ids:
        u = db.get(User, uid)
        if u and u.status == "pending":
            u.status = "active"
            u.rejection_note = None
            count += 1
    db.commit()
    return {"message": f"تم قبول {count} طلب"}


@router.post("/bulk/reject")
def bulk_reject(
    data: BulkUserIds,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    count = 0
    for uid in data.user_ids:
        u = db.get(User, uid)
        if u and u.status == "pending":
            u.status = "rejected"
            count += 1
    db.commit()
    return {"message": f"تم رفض {count} طلب"}


@router.post("/bulk/activate")
def bulk_activate(
    data: BulkUserIds,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    count = 0
    for uid in data.user_ids:
        u = db.get(User, uid)
        if u and u.status in ("rejected", "withdrawn"):
            u.status = "active"
            count += 1
    db.commit()
    return {"message": f"تم تفعيل {count} مشارك"}


@router.post("/bulk/withdraw")
def bulk_withdraw(
    data: BulkUserIds,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    count = 0
    for uid in data.user_ids:
        u = db.get(User, uid)
        if u and u.status == "active":
            u.status = "withdrawn"
            count += 1
    db.commit()
    return {"message": f"تم سحب {count} مشارك"}


@router.post("/bulk/assign-halqa")
def bulk_assign_halqa(
    data: BulkAssignHalqa,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    count = 0
    for uid in data.user_ids:
        u = db.get(User, uid)
        if u:
            u.halqa_id = data.halqa_id
            count += 1
    db.commit()
    return {"message": f"تم تعيين الحلقة لـ {count} مشارك"}


# ─── Users Export ─────────────────────────────────────────────────────────────


@router.get("/export-users")
def export_users(
    format: str = Query("xlsx"),
    status: str = Query(None),
    gender: str = Query(None),
    halqa_id: int = Query(None),
    search: str = Query(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Export users list with all personal info."""
    import csv
    from openpyxl import Workbook

    query = db.query(User)
    if status:
        query = query.filter_by(status=status)
    if gender:
        male_values = {"male", "ذكر"}
        female_values = {"female", "أنثى"}
        match_set = male_values if gender == "male" else female_values
        query = query.filter(User.gender.in_(match_set))
    if halqa_id:
        query = query.filter_by(halqa_id=int(halqa_id))
    if search:
        query = query.filter(
            or_(User.full_name.ilike(f"%{search}%"), User.email.ilike(f"%{search}%"))
        )

    users_list = query.order_by(User.created_at.desc()).all()

    gender_map = {"male": "ذكر", "female": "أنثى"}
    status_map = {"active": "نشط", "pending": "قيد المراجعة", "rejected": "مرفوض", "withdrawn": "منسحب"}
    role_map = {"participant": "مشارك", "supervisor": "مشرف", "super_admin": "سوبر آدمن"}

    rows = []
    for u in users_list:
        rows.append({
            "رقم العضوية": u.member_id,
            "الاسم": u.full_name,
            "الجنس": gender_map.get(u.gender, u.gender),
            "العمر": u.age,
            "الهاتف": u.phone,
            "البريد": u.email,
            "الدولة": u.country,
            "الحالة": status_map.get(u.status, u.status),
            "الصلاحية": role_map.get(u.role, u.role),
            "الحلقة": u.halqa.name if u.halqa else "-",
            "المشرف": u.halqa.supervisor.full_name if u.halqa and u.halqa.supervisor else "-",
            "مصدر التسجيل": u.referral_source or "",
            "تاريخ التسجيل": u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
        })

    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "المستخدمون"
        ws.sheet_view.rightToLeft = True
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row[h] for h in headers])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=users_report.xlsx"},
        )
    else:
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        csv_bytes = "\ufeff" + output.getvalue()
        return Response(
            content=csv_bytes.encode("utf-8"),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": "attachment; filename=users_report.csv"},
        )
